"""
Upload si parsare fisiere Excel/CSV
Detecteaza automat structura si mapeaza la indicatorii din dashboard.
"""

import pandas as pd
import io
import streamlit as st
from datetime import datetime


# ── Parsare fisier ────────────────────────────────────────────────────────────

def parse_uploaded_file(uploaded_file) -> dict:
    """
    Parseaza un fisier Excel sau CSV incarcat via st.file_uploader.
    Returneaza dict cu: {sheet_name: DataFrame}
    """
    name = uploaded_file.name.lower()
    result = {}

    try:
        if name.endswith(".csv"):
            # Detecteaza separatorul automat
            content = uploaded_file.read().decode("utf-8-sig", errors="replace")
            sep = ";" if content.count(";") > content.count(",") else ","
            df = pd.read_csv(io.StringIO(content), sep=sep,
                             thousands=".", decimal=",")
            df = _clean_df(df)
            result["Sheet1"] = df

        elif name.endswith((".xlsx", ".xls")):
            xls = pd.ExcelFile(uploaded_file)
            for sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet,
                                   header=None)  # citim fara header initial
                df = _detect_header(df)
                df = _clean_df(df)
                result[sheet] = df

    except Exception as e:
        st.error(f"Eroare la parsarea fisierului: {e}")

    return result


def _detect_header(df: pd.DataFrame) -> pd.DataFrame:
    """Detecteaza randul de header si il seteaza ca index de coloane."""
    # Cauta primul rand cu mai mult de 50% celule ne-goale
    for i, row in df.iterrows():
        filled = row.notna().sum()
        if filled >= max(2, len(df.columns) * 0.4):
            df.columns = [str(c).strip() for c in df.iloc[i]]
            df = df.iloc[i+1:].reset_index(drop=True)
            return df
    return df


def _clean_df(df: pd.DataFrame) -> pd.DataFrame:
    """Curata DataFrame: strip, converteste numere, elimina randuri goale."""
    df = df.dropna(how="all").dropna(axis=1, how="all")
    df.columns = [str(c).strip() for c in df.columns]

    for col in df.columns:
        # Incearca conversie numerica
        try:
            cleaned = df[col].astype(str).str.replace(" ", "").str.replace(",", ".").str.replace("%","")
            numeric = pd.to_numeric(cleaned, errors="coerce")
            if numeric.notna().sum() / len(numeric) > 0.5:
                df[col] = numeric
        except Exception:
            pass

        # Incearca conversie data
        if df[col].dtype == object:
            try:
                df[col] = pd.to_datetime(df[col], dayfirst=True, errors="ignore")
            except Exception:
                pass

    return df


# ── Detectare tip serie ───────────────────────────────────────────────────────

INDICATOR_KEYWORDS = {
    "pib":        ["pib", "gdp", "produs intern brut"],
    "ipc":        ["ipc", "inflatie", "preturi consum", "cpi"],
    "export":     ["export", "exporturi"],
    "import":     ["import", "importuri"],
    "somaj":      ["somaj", "someri", "somer", "unemployment"],
    "salariu":    ["salariu", "salarii", "castig salarial", "wage"],
    "pensie":     ["pensie", "pensii", "pension"],
    "curs":       ["curs", "valutar", "eur", "usd", "exchange"],
    "rezerve":    ["rezerve", "reserve"],
    "buget":      ["buget", "fiscal", "venituri", "cheltuieli"],
}


def detect_indicator_type(col_name: str) -> str:
    """Detecteaza tipul de indicator dintr-un nume de coloana."""
    col_lower = col_name.lower()
    for tip, keywords in INDICATOR_KEYWORDS.items():
        if any(kw in col_lower for kw in keywords):
            return tip
    return "necunoscut"


def auto_map_columns(df: pd.DataFrame) -> dict:
    """
    Mapeaza automat coloanele unui DataFrame la indicatorii din dashboard.
    Returneaza dict {coloana: tip_indicator}
    """
    mapping = {}
    for col in df.columns:
        tip = detect_indicator_type(col)
        if tip != "necunoscut":
            mapping[col] = tip
    return mapping


# ── Template Excel de import ──────────────────────────────────────────────────

def generate_template() -> bytes:
    """
    Genereaza un fisier Excel template cu structura standardizata
    pentru importul datelor in dashboard.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Sheet 1: Sector Real ──
    ws1 = wb.active
    ws1.title = "Sector Real"

    headers = ["An", "PIB nominal (mil. MDL)", "Crestere PIB real (%)",
               "PIB/locuitor (USD PPP)", "FBCF (% PIB)"]
    sample  = [
        [2020, 196300, -7.4, 11800, 22.1],
        [2021, 279100, 13.9, 14820, 23.4],
        [2022, 338120, -5.0, 15640, 22.8],
        [2023, 370840, -5.9, 16200, 21.8],
        [2024, 388600,  2.1, 18200, 21.3],
    ]

    _write_sheet(ws1, headers, sample, "Sector Real — PIB si crestere economica")

    # ── Sheet 2: Preturi ──
    ws2 = wb.create_sheet("Preturi")
    headers2 = ["Luna/An", "IPC total (%)", "Inflatie de baza (%)",
                "Alimente (%)", "Energie (%)"]
    sample2  = [
        ["Ian-24", 6.0, 4.5, 7.2, 4.1],
        ["Feb-24", 5.8, 4.3, 7.0, 3.9],
        ["Mar-24", 5.5, 4.2, 6.8, 3.7],
    ]
    _write_sheet(ws2, headers2, sample2, "Preturi — IPC si componente")

    # ── Sheet 3: Sector Extern ──
    ws3 = wb.create_sheet("Sector Extern")
    headers3 = ["An", "Export (mil. USD)", "Import (mil. USD)",
                "Deficit CA (% PIB)", "Rezerve BNM (mil. USD)"]
    sample3  = [
        [2022, 3188, 7562, -11.2, 3960],
        [2023, 3284, 7648,  -9.8, 4210],
        [2024, 3421, 7810,  -8.4, 4450],
    ]
    _write_sheet(ws3, headers3, sample3, "Sector Extern — comert si balanta de plati")

    # ── Sheet 4: Sector Monetar ──
    ws4 = wb.create_sheet("Sector Monetar")
    headers4 = ["Data", "Rata BNM (%)", "Curs EUR/MDL", "Curs USD/MDL", "M3 (mil. MDL)"]
    sample4  = [
        ["2024-01-04", 7.5, 19.18, 17.96, 82400],
        ["2024-04-04", 6.0, 19.35, 18.10, 84200],
        ["2024-07-04", 4.5, 19.40, 18.15, 86800],
    ]
    _write_sheet(ws4, headers4, sample4, "Sector Monetar — politica monetara")

    # ── Sheet 5: Sector Public ──
    ws5 = wb.create_sheet("Sector Public")
    headers5 = ["An", "Venituri (% PIB)", "Cheltuieli (% PIB)",
                "Deficit (% PIB)", "Datorie publica (% PIB)"]
    sample5  = [
        [2022, 33.2, 36.2, -3.0, 33.6],
        [2023, 32.7, 35.5, -2.8, 32.1],
        [2024, 33.1, 36.5, -3.4, 31.2],
    ]
    _write_sheet(ws5, headers5, sample5, "Sector Public — finante publice")

    # ── Sheet 6: Sector Social ──
    ws6 = wb.create_sheet("Sector Social")
    headers6 = ["An", "Rata somajului (%)", "Salariu mediu (MDL)",
                "Pensie medie (MDL)", "Rata saraciei (%)"]
    sample6  = [
        [2022, 3.1, 8642, 2460, 23.8],
        [2023, 2.8, 9850, 2840, 21.2],
        [2024, 2.6, 11200, 3200, 19.8],
    ]
    _write_sheet(ws6, headers6, sample6, "Sector Social — piata muncii si saracie")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_sheet(ws, headers, data, title):
    """Helper pentru formatarea unui sheet Excel (numar variabil de coloane)."""
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    n = max(len(headers), 1)
    end_col = get_column_letter(n)
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=11)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.font = Font(bold=True, size=10)

    for i, row in enumerate(data, 3):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)

    for col in ws.columns:
        cells = [c for c in col if hasattr(c, "column_letter")]
        if not cells:
            continue
        max_len = max(len(str(c.value or "")) for c in cells)
        ws.column_dimensions[cells[0].column_letter].width = min(max_len + 4, 30)


def _safe_rows(df: pd.DataFrame) -> list:
    """Converteste DataFrame la lista de liste cu tipuri Python native (pentru openpyxl)."""
    rows = []
    for _, row in df.iterrows():
        r = []
        for v in row:
            try:
                is_na = pd.isna(v)
            except Exception:
                is_na = False
            if is_na:
                r.append(None)
            elif hasattr(v, "item"):
                r.append(v.item())
            else:
                r.append(v)
        rows.append(r)
    return rows


def generate_bns_export() -> bytes:
    """
    Exporta datele BNS afisate in dashboard intr-un Excel cu sheet-uri per sector:
    PIB, Export/Import lunar, Industrie, Salarii & Piata muncii,
    Venituri bugetare, Cheltuieli bugetare, Monetar.
    """
    import openpyxl
    import os

    wb_file = openpyxl.Workbook()
    sheets_added = []

    _here = os.path.dirname(os.path.abspath(__file__))
    _root = os.path.dirname(_here)

    def _find(fname):
        for p in [os.path.join(_root, "data", fname),
                  os.path.join("data", fname), fname]:
            if os.path.exists(p):
                return p
        return None

    def _add(title, df, sheet_title):
        if df is None or df.empty:
            return
        if not sheets_added:
            ws = wb_file.active
            ws.title = title
        else:
            ws = wb_file.create_sheet(title)
        _write_sheet(ws, list(df.columns), _safe_rows(df), sheet_title)
        sheets_added.append(title)

    # ── 1. PIB & Sector Real ──────────────────────────────────────────────────
    real_path = _find("Real.xlsx")
    if real_path:
        try:
            df_pib = pd.read_excel(real_path, sheet_name="PIB")
            df_pib["An"] = pd.to_numeric(df_pib["An"], errors="coerce")
            df_pib = df_pib.dropna(subset=["An"]).sort_values("An").reset_index(drop=True)
            df_pib["An"] = df_pib["An"].astype(int)
            _add("PIB — Sector Real", df_pib, "PIB si crestere economica — Sursa: BNS / Real.xlsx")
        except Exception:
            pass

    # ── 2. Export si Import lunar (BNS EXT015000) ─────────────────────────────
    try:
        from utils.api_comert_extern import get_comert_ext_bns
        df_ext = get_comert_ext_bns()["data"]
        if not df_ext.empty:
            df_total = df_ext[df_ext["grupa"] == "Total"].copy()
            keep = [c for c in ["an", "label", "export_mil_usd", "import_mil_usd", "sold_mil_usd"]
                    if c in df_total.columns]
            df_total = df_total[keep].rename(columns={
                "an": "An", "label": "Perioada",
                "export_mil_usd": "Export (mil. USD)",
                "import_mil_usd": "Import (mil. USD)",
                "sold_mil_usd":   "Sold comercial (mil. USD)",
            })
            _add("Export si Import (lunar)", df_total,
                 "Export, Import, Sold comercial — BNS EXT015000 (lunar, mil. USD)")
    except Exception:
        pass

    # ── 3. Industrie ─────────────────────────────────────────────────────────
    if real_path:
        try:
            df_ind = pd.read_excel(real_path, sheet_name="Industrie")
            df_ind["An"] = pd.to_numeric(df_ind["An"], errors="coerce")
            df_ind = df_ind.dropna(subset=["An"]).sort_values("An").reset_index(drop=True)
            df_ind["An"] = df_ind["An"].astype(int)
            for col in df_ind.columns:
                if col != "An":
                    df_ind[col] = pd.to_numeric(df_ind[col], errors="coerce")
            _add("Industrie (BNS)", df_ind, "Productie industriala — Sursa: BNS / Real.xlsx")
        except Exception:
            pass

    # ── 4. Salarii si Piata Muncii ────────────────────────────────────────────
    try:
        from utils.api_social import get_salarii_anual, get_piata_muncii_anual
        df_sal = get_salarii_anual()["data"]
        df_pm  = get_piata_muncii_anual()["data"]
        if not df_sal.empty:
            if not df_pm.empty:
                df_soc = df_sal.merge(df_pm, on="an", how="outer").sort_values("an").reset_index(drop=True)
            else:
                df_soc = df_sal
            df_soc = df_soc.rename(columns={
                "an":              "An",
                "total":           "Salariu mediu brut (MDL)",
                "rata_somaj":      "Rata somajului (%)",
                "rata_ocupare":    "Rata ocuparii (%)",
                "rata_activitate": "Rata activitatii (%)",
            })
            _add("Salarii si Piata Muncii", df_soc, "Salarii si piata muncii — Sursa: BNS API")
    except Exception:
        pass

    # ── 5. Finante Publice ────────────────────────────────────────────────────
    bpn_path = _find("Date_Public_BPN.xlsx")
    if bpn_path:
        for sheet_src, ws_name, ws_title in [
            ("Venituri",   "Venituri Bugetare",  "Venituri bugetare (BPN) — Sursa: BNS"),
            ("Cheltuieli", "Cheltuieli Bugetare", "Cheltuieli bugetare (BPN) — Sursa: BNS"),
        ]:
            try:
                df_bp = pd.read_excel(bpn_path, sheet_name=sheet_src, index_col=0)
                df_bp = df_bp.T.reset_index().rename(columns={"index": "An"})
                df_bp["An"] = pd.to_numeric(df_bp["An"], errors="coerce")
                df_bp = df_bp.dropna(subset=["An"]).sort_values("An").reset_index(drop=True)
                df_bp["An"] = df_bp["An"].astype(int)
                _add(ws_name, df_bp, ws_title)
            except Exception:
                pass

    # ── 6. Monetar ────────────────────────────────────────────────────────────
    try:
        from data.excel_loader import load_monetar
        df_mon = load_monetar()["data"]
        if not df_mon.empty:
            if "an" in df_mon.columns:
                df_mon = df_mon.rename(columns={"an": "An"})
            _add("Monetar (BNM)", df_mon, "Date monetare — Date_Sector_Monetar.xlsx / BNM")
    except Exception:
        pass

    if not sheets_added:
        ws_def = wb_file.active
        ws_def.title = "Info"
        ws_def["A1"] = "Date indisponibile — verificati fisierele din /data/"

    buf = io.BytesIO()
    wb_file.save(buf)
    buf.seek(0)
    return buf.read()


def generate_data_export() -> bytes:
    """
    Exporta date reale din IMF si World Bank intr-un Excel cu mai multe sheet-uri:
    WEO Moldova (PIB, export, import, inflatie, datorie, cont curent, somaj),
    World Bank (export/import % PIB, remitente, saracie, populatie),
    comparativ regional PIB si inflatie.
    """
    import openpyxl
    from utils.api_imf import imf_weo_multi, wb_multi, imf_regional_compare

    wb_file = openpyxl.Workbook()

    # ── Sheet 1: IMF WEO Moldova ──────────────────────────────────────────────
    ws1 = wb_file.active
    ws1.title = "IMF — WEO Moldova"
    df_imf = imf_weo_multi(
        ["NGDP_RPCH", "NGDPD", "PCPIPCH", "TX_RPCH", "TM_RPCH",
         "BCA_NGDPD", "GGXWDG_NGDP", "LUR"],
        ani_start=2015
    )
    if not df_imf.empty:
        df_imf = df_imf.rename(columns={
            "an":           "An",
            "NGDP_RPCH":   "Crestere PIB real (%)",
            "NGDPD":       "PIB nominal (mld. USD)",
            "PCPIPCH":     "Inflatie IPC (%)",
            "TX_RPCH":     "Crestere exporturi vol. (%)",
            "TM_RPCH":     "Crestere importuri vol. (%)",
            "BCA_NGDPD":   "Cont curent (% PIB)",
            "GGXWDG_NGDP": "Datorie publica (% PIB)",
            "LUR":         "Rata somajului (%)",
        }).round(2)
        _write_sheet(ws1, list(df_imf.columns), _safe_rows(df_imf),
                     "IMF WEO — PIB, Export, Import, Inflatie, Datorie, Moldova (2015–prezent+prognoze)")

    # ── Sheet 2: World Bank Moldova ───────────────────────────────────────────
    ws2 = wb_file.create_sheet("World Bank — Moldova")
    df_wb = wb_multi({
        "PIB nominal (USD)":              "NY.GDP.MKTP.CD",
        "PIB/locuitor (USD)":             "NY.GDP.PCAP.CD",
        "Export bunuri+servicii (% PIB)": "NE.EXP.GNFS.ZS",
        "Import bunuri+servicii (% PIB)": "NE.IMP.GNFS.ZS",
        "Remitente (% PIB)":              "BX.TRF.PWKR.DT.GD.ZS",
        "Saracie (%)":                    "SI.POV.NAHC",
        "Somaj (% forta munca)":          "SL.UEM.TOTL.ZS",
        "Populatie (locuitori)":          "SP.POP.TOTL",
        "Datorie publica (% PIB)":        "GC.DOD.TOTL.GD.ZS",
    }, ani_start=2010)
    if not df_wb.empty:
        df_wb = df_wb.rename(columns={"an": "An"}).round(2)
        _write_sheet(ws2, list(df_wb.columns), _safe_rows(df_wb),
                     "World Bank — Export, Import, Remitente, Saracie, Moldova (2010–prezent)")

    # ── Sheet 3: IMF comparativ regional PIB ─────────────────────────────────
    ws3 = wb_file.create_sheet("IMF — PIB regional")
    df_reg = imf_regional_compare("NGDP_RPCH")
    if not df_reg.empty:
        df_reg = df_reg.rename(columns={"an": "An"}).round(2)
        _write_sheet(ws3, list(df_reg.columns), _safe_rows(df_reg),
                     "IMF WEO — Crestere PIB real (%) comparativ regional")

    # ── Sheet 4: IMF comparativ regional inflatie ─────────────────────────────
    ws4 = wb_file.create_sheet("IMF — Inflatie regionala")
    df_inf = imf_regional_compare("PCPIPCH")
    if not df_inf.empty:
        df_inf = df_inf.rename(columns={"an": "An"}).round(2)
        _write_sheet(ws4, list(df_inf.columns), _safe_rows(df_inf),
                     "IMF WEO — Inflatie IPC (%) comparativ regional")

    buf = io.BytesIO()
    wb_file.save(buf)
    buf.seek(0)
    return buf.read()
